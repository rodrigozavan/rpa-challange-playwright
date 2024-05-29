from typing import List
import openpyxl


def get_header(file_path) -> List[str]:
    """
    Retorna o cabeçalho de um arquivo Excel especificado.

    Args:
        file_path (str): O caminho do arquivo Excel.

    Returns:
        list: Uma lista contendo os cabeçalhos do arquivo Excel.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header = [
        cell.value for cell in ws[1][0:-1]
    ]
    
    return header
    

def read_excel(file_path, min_row=2, max_col=None) -> List[dict]:
    """
    Lê um arquivo Excel e retorna os dados contidos nele.

    Args:
        file_path (str): O caminho do arquivo Excel a ser lido.
        min_row (int, optional): A linha mínima a partir da qual os dados serão lidos. O padrão é 2.
        max_col (int, optional): A coluna máxima até a qual os dados serão lidos. Se não for fornecido, será lido até a penultima coluna.

    Returns:
        list: Uma lista contendo os dados lidos do arquivo Excel.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data = []
    
    if max_col is None:
        max_col = ws.max_column - 1
        
    header = get_header(file_path)
    
    for row in ws.iter_rows(values_only=True, min_row=min_row, max_col=max_col):
        if not all([cell is None for cell in row]):
            data_row = {}
            for index, item in enumerate(row):
                data_row[header[index]] = item
            
            data.append(data_row)
            
    return data
